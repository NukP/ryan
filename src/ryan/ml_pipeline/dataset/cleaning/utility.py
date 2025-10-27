"""This modeule contains utility functions that are used with analyzing data/metadata."""

import os
import pandas as pd
from pandas import DataFrame
from typing import List, Optional
import shutil
import zipfile
import json
from collections import Counter
from datetime import datetime
import tempfile
from openpyxl import load_workbook
from pathlib import Path
from yadg.extractors.fusion import json as fusion_json


def load_dataframes(directory: str) -> dict:
    """
    Load Excel files from a specified directory into a dictionary of dataframes.

    Args:
        directory (str): The path to the directory containing the Excel files.

    Returns:
        dict: A dictionary with filenames as keys and corresponding dataframes as values.

    Raises:
        ValueError: If there is an error in loading dataframes.
    """
    try:
        file_paths = [os.path.join(directory, file) for file in os.listdir(directory) if file.endswith(".xlsx")]
        dataframes = {file: pd.read_excel(file_path) for file, file_path in zip(os.listdir(directory), file_paths)}
        return dataframes
    except Exception as e:
        raise ValueError(f"Error loading dataframes: {e}")


def get_value_from_df(df: DataFrame, input_row_name: str, input_column: str = "Metadata", output_column_name: str = "Value") -> any:
    """
    Retrieve a value from a DataFrame based on a specified row and column.

    This function searches for a row in the DataFrame where the value in `input_column`
    matches `input_row_name`. If a matching row is found, it returns the value from the
    `output_column_name`. If no matching row is found, a `ValueError` is raised.

    Parameters:
    -----------
    df : DataFrame
        The DataFrame to search within.
    input_row_name : str
        The value to search for in the `input_column`.
    input_column : str, optional
        The name of the column where the function will search for `input_row_name`.
        Default is 'Metadata'.
    output_column_name : str, optional
        The name of the column from which to retrieve the value.
        Default is "Value".

    Returns:
    --------
    any
        The value found in the `output_column_name` for the row where `input_column` matches `input_row_name`.

    Raises:
    -------
    ValueError
        If no row with `input_column` equal to `input_row_name` is found in the DataFrame.
    """
    target_row = df[df[input_column] == input_row_name]
    if not target_row.empty:
        return target_row[output_column_name].values[0]
    else:
        raise ValueError(f"Row with {input_column} = {input_row_name} not found in the DataFrame.")


def convert_to_dict(string: str, name_1: str, name_2: str, name_3: str) -> dict[str, str]:
    """
    Converts a string in the format "[x,y,z]" into a dictionary with specified key names.

    Parameters:
    string (str): The input string in the format "[x,y,z]".
    name_1 (str): The key name for the first value (x).
    name_2 (str): The key name for the second value (y).
    name_3 (str): The key name for the third value (z).

    Returns:
    dict: A dictionary with the specified key names and corresponding values from the string.

    Example:
    >>> result = convert_to_dict("[1,2,3]", "name_1", "name_2", "name_3")
    >>> print(result)
    {'name_1': '1', 'name_2': '2', 'name_3': '3'}
    """
    # Remove the square brackets and split the string by commas
    values = string.strip("[]").split(",")

    # Create the dictionary using the provided key names and extracted values
    return {name_1: values[0], name_2: values[1], name_3: values[2]}


def print_list_variety(values: list) -> None:
    """
    Prints the variety of unique values in the list and their corresponding frequencies.

    Parameters:
    values (list): A list containing values of any data type.

    Example:
    >>> print_variety_and_frequency([1, 2, 2, 3, 3, 3])
    Value: 1, Frequency: 1
    Value: 2, Frequency: 2
    Value: 3, Frequency: 3
    """
    # Use Counter to count the frequencies of each unique value
    frequency_count = Counter(values)

    # Print the value and its frequency
    for value, frequency in frequency_count.items():
        print(f"Value: {value}, Frequency: {frequency}")


def get_value_from_excel(
    excel_path: str, input_row_name: str, input_column: str = "Metadata", output_column_name: str = "Value", sheet_name: str = "Metadata"
) -> any:
    """
    Retrieve a value from an Excel file based on a specified row and column.

    This function reads the specified sheet from the Excel file into a DataFrame,
    searches for a row where the value in `input_column` matches `input_row_name`,
    and returns the value from the `output_column_name`.

    Parameters:
    -----------
    excel_path : str
        Path to the Excel file.
    input_row_name : str
        The value to search for in the `input_column`.
    input_column : str, optional
        The name of the column where the function will search for `input_row_name`.
        Default is 'Metadata'.
    output_column_name : str, optional
        The name of the column from which to retrieve the value.
        Default is "Value".
    sheet_name : str, optional
        The name of the sheet to read. Default is "Metadata".

    Returns:
    --------
    any
        The value found in the `output_column_name` for the row where `input_column` matches `input_row_name`.

    Raises:
    -------
    ValueError
        If no row with `input_column` equal to `input_row_name` is found in the DataFrame.
    """
    # Load the Excel file into a DataFrame
    df = pd.read_excel(excel_path, sheet_name=sheet_name)

    # Call the original function with the DataFrame
    return get_value_from_df(df, input_row_name, input_column, output_column_name)


def set_value_in_excel(
    excel_path: str,
    input_row_name: str,
    new_value: any,
    input_column: str = "Metadata",
    output_column_name: str = "Value",
    sheet_name: str = "Metadata",
) -> None:
    """
    Update a specific value in an Excel file while preserving formatting and structure.

    This function searches for a row in the specified sheet of the Excel file where the value
    in the `input_column` matches `input_row_name`, and updates the value in the `output_column_name`
    to `new_value`. It preserves all formatting and structure of the Excel file.

    Parameters:
    -----------
    excel_path : str
        Path to the Excel file.
    input_row_name : str
        The value to search for in the `input_column`.
    new_value : any
        The new value to assign to the `output_column_name` of the matching row.
    input_column : str, optional
        The name of the column where the function will search for `input_row_name`.
        Default is 'Metadata'.
    output_column_name : str, optional
        The name of the column where the value will be updated.
        Default is "Value".
    sheet_name : str, optional
        The name of the sheet to read and modify. Default is "Metadata".

    Returns:
    --------
    None

    Raises:
    -------
    ValueError
        If no row with `input_column` equal to `input_row_name` is found in the sheet.
    """
    # Load the workbook and sheet
    workbook = load_workbook(excel_path)
    sheet = workbook[sheet_name]

    # Find the input and output column indices
    input_col_index = None
    output_col_index = None
    for col_idx, cell in enumerate(sheet[1], start=1):  # Assuming the first row contains headers
        if cell.value == input_column:
            input_col_index = col_idx
        if cell.value == output_column_name:
            output_col_index = col_idx
    if input_col_index is None or output_col_index is None:
        raise ValueError(f"Columns '{input_column}' or '{output_column_name}' not found in the sheet.")

    # Search for the row with the matching input_row_name
    target_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Start from the second row
        if row[input_col_index - 1].value == input_row_name:
            target_row = row_idx
            break
    if target_row is None:
        raise ValueError(f"Row with {input_column} = {input_row_name} not found in the sheet.")

    # Update the value in the output column
    sheet.cell(row=target_row, column=output_col_index, value=new_value)

    # Save the workbook back to the original file
    workbook.save(excel_path)


def delete_metadata_files(path: str) -> None:
    """
    Deletes all files ending with '-metadata.xlsx' in the specified directory and its subdirectories.

    Args:
        path (str): The root directory to start searching for files to delete.

    Returns:
        None
    """
    if not os.path.isdir(path):
        raise ValueError(f"The provided path '{path}' is not a valid directory.")

    for root, _, files in os.walk(path):
        for file in files:
            if file.endswith("-metadata.xlsx"):
                file_path = os.path.join(root, file)
                try:
                    os.remove(file_path)
                except OSError as e:
                    print(f"Error deleting file {file_path}: {e}")


def check_metadata_availability(folder_dir: str) -> None:
    """
    Checks the availability of metadata files in the given folder path.

    Args:
        folder_dir (str): The root folder directory to start the search.

    Returns:
        None
    """
    if not os.path.isdir(folder_dir):
        raise ValueError(f"The provided path '{folder_dir}' is not a valid directory.")

    for folder_name in os.listdir(folder_dir):
        # Skip temporary or hidden files/folders
        if folder_name.startswith("~"):
            continue

        folder_path = os.path.join(folder_dir, folder_name)

        if os.path.isdir(folder_path):  # Ensure it's a directory
            if folder_name.startswith("Multiplex"):
                # Handle "Multiplex" folders
                for sub_folder_name in os.listdir(folder_path):
                    if sub_folder_name.startswith("U"):
                        sub_folder_path = os.path.join(folder_path, sub_folder_name)
                        if os.path.isdir(sub_folder_path):
                            # Check for metadata files in the sub-folder
                            metadata_files = [
                                f for f in os.listdir(sub_folder_path) if f.endswith("metadata.xlsx") and not f.startswith("~")
                            ]
                            if not metadata_files:
                                print(f"Folder: {folder_name}, unit: {sub_folder_name} is missing metadata.")
            else:
                # Handle non-"Multiplex" folders
                metadata_files = [f for f in os.listdir(folder_path) if f.endswith("metadata.xlsx") and not f.startswith("~")]
                if not metadata_files:
                    print(f"Folder: {folder_name} is missing metadata.")


def push_metadata_files(data_folder: str, metadata_folder: str, verbose: bool = False, show_intend_path: bool = False) -> None:
    """
    Copies metadata files from the metadata folder into the correct directories in the data folder. Handles
    inconsistencies between underscores and hyphens in file and folder names for matching. Dynamically tries adding
    "Multiplex_" or "MultiplexN_" prefixes for files with "UX-metadata.xlsx" patterns.

    Args:
        data_folder (str): Path to the folder containing experimental data.
        metadata_folder (str): Path to the folder containing metadata files.
        verbose (bool): If True, prints successful file copy messages. Defaults to False.
        show_intend_path (bool): If True, prints debugging information for intended paths. Defaults to False.

    Returns:
        None
    """
    if not os.path.isdir(data_folder):
        raise ValueError(f"The provided data folder path '{data_folder}' is not a valid directory.")
    if not os.path.isdir(metadata_folder):
        raise ValueError(f"The provided metadata folder path '{metadata_folder}' is not a valid directory.")

    def normalize_name(name: str) -> str:
        """
        Normalizes a name by replacing underscores and hyphens with spaces for comparison.

        Args:
            name (str): The name to normalize.

        Returns:
            str: The normalized name.
        """
        return name.replace("_", " ").replace("-", " ").lower()

    metadata_files = [file for file in os.listdir(metadata_folder) if file.endswith("-metadata.xlsx")]

    for metadata_file in metadata_files:
        metadata_src = os.path.join(metadata_folder, metadata_file)

        if "-U" in metadata_file:  # Detect UX-metadata.xlsx pattern
            parts = metadata_file.split("-")
            base_name = "-".join(parts[:-2])  # Everything except UX and metadata.xlsx
            unit = parts[-2]  # UX part

            # Try to find the folder directly
            target_folder = None
            for folder in os.listdir(data_folder):
                if normalize_name(base_name) in normalize_name(folder):
                    target_folder = os.path.join(data_folder, folder, unit)
                    break

            # If no match, try adding Multiplex_ or MultiplexN_ prefixes
            if not target_folder or not os.path.isdir(target_folder):
                for prefix in ["Multiplex_"] + [f"Multiplex{n}_" for n in range(2, 9)]:
                    for folder in os.listdir(data_folder):
                        prefixed_name = prefix + base_name
                        if normalize_name(prefixed_name) in normalize_name(folder):
                            target_folder = os.path.join(data_folder, folder, unit)
                            break
                    if target_folder and os.path.isdir(target_folder):
                        break

        else:
            # Handle non-"Multiplex" metadata files
            base_name = metadata_file.replace("-metadata.xlsx", "")
            target_folder = None
            for folder in os.listdir(data_folder):
                if normalize_name(base_name) == normalize_name(folder):
                    target_folder = os.path.join(data_folder, folder)
                    break

        # Debugging: show intended path if enabled
        if show_intend_path:
            print(f"Metadata: {metadata_file}, Intended Path: {target_folder if target_folder else 'No match found'}")

        if target_folder and os.path.isdir(target_folder):
            try:
                shutil.copy(metadata_src, os.path.join(target_folder, metadata_file))
                if verbose:
                    print(f"Copied: {metadata_file} to {target_folder}")
            except Exception as e:
                print(f"Failed to copy {metadata_file} to {target_folder}: {e}")
        else:
            print(f"Target folder does not exist for metadata file: {metadata_file}")
            if not show_intend_path:
                print(f"Intended Path: {target_folder if target_folder else 'No match found'}")


def clean_point_dir(folder_path: str) -> None:
    """
    Recursively renames files and folders in a directory and its subdirectories. Replaces occurrences of "22.5" or
    "22p5" with "22point5" in their names.

    Args:
        folder_path (str): Path to the folder to process.

    Returns:
        None
    """
    if not os.path.isdir(folder_path):
        raise ValueError(f"The provided path '{folder_path}' is not a valid directory.")

    # Traverse the directory tree, renaming files and folders
    for root, dirs, files in os.walk(folder_path, topdown=False):  # Process inner folders first
        # Rename files
        for file in files:
            if "22.5" in file or "22p5" in file:
                old_file_path = os.path.join(root, file)
                new_file_name = file.replace("22.5", "22point5").replace("22p5", "22point5")
                new_file_path = os.path.join(root, new_file_name)

                try:
                    os.rename(old_file_path, new_file_path)
                    print(f"Renamed File: {old_file_path} -> {new_file_path}")
                except Exception as e:
                    print(f"Failed to rename file {old_file_path}: {e}")

        # Rename folders
        for dir_name in dirs:
            if "22.5" in dir_name or "22p5" in dir_name:
                old_dir_path = os.path.join(root, dir_name)
                new_dir_name = dir_name.replace("22.5", "22point5").replace("22p5", "22point5")
                new_dir_path = os.path.join(root, new_dir_name)

                try:
                    os.rename(old_dir_path, new_dir_path)
                    print(f"Renamed Folder: {old_dir_path} -> {new_dir_path}")
                except Exception as e:
                    print(f"Failed to rename folder {old_dir_path}: {e}")


def pull_metadata_files(dir_data: str, dir_transfer_to: str, to_move: bool) -> None:
    """
    Copies or moves metadata files (-metadata.xlsx) from a data folder to a specified directory.

    Args:
        dir_data (str): The directory to search for metadata files.
        dir_transfer_to (str): The directory to transfer the metadata files to.
        to_move (bool): If True, moves the files; if False, copies them.

    Returns:
        None
    """
    if not os.path.isdir(dir_data):
        raise ValueError(f"The provided dir_data path '{dir_data}' is not a valid directory.")

    # Create the transfer-to directory if it doesn't exist
    os.makedirs(dir_transfer_to, exist_ok=True)

    # Traverse all folders and subfolders in dir_data
    for root, _, files in os.walk(dir_data):
        for file in files:
            if file.endswith("-metadata.xlsx"):  # Check if the file is a metadata file
                src_path = os.path.join(root, file)
                dst_path = os.path.join(dir_transfer_to, file)

                try:
                    if to_move:
                        shutil.move(src_path, dst_path)
                        print(f"Moved: {src_path} -> {dst_path}")
                    else:
                        shutil.copy(src_path, dst_path)
                        print(f"Copied: {src_path} -> {dst_path}")
                except Exception as e:
                    print(f"Failed to {'move' if to_move else 'copy'} {src_path}: {e}")


def check_gc_data(dir_output: str) -> None:
    """
    Checks if each subfolder in the specified directory contains a file ending with '.GCdata.xlsx'.

    Args:
        dir_output (str): The path to the directory containing subfolders.

    Returns:
        None
    """
    # List all folders in the specified directory
    folders = [folder for folder in os.listdir(dir_output) if os.path.isdir(os.path.join(dir_output, folder))]

    # Iterate through each folder
    for folder in folders:
        folder_path = os.path.join(dir_output, folder)
        # Check if any file in the folder ends with '.GCdata.xlsx'
        gc_data_exists = any(file.endswith(".GCdata.xlsx") for file in os.listdir(folder_path))

        if not gc_data_exists:
            print(f"Folder {folder} does not contain GC data.")


def analyze_annotations(data_folder_dir: str) -> None:
    """
    Analyze and print the occurrences of unique values in "annotations:name" and "sequence:location" within the first
    zip file ending with "-GC.zip" found in the given directory.

    Args:
        data_folder_dir (str): Directory containing the data folder.
    """
    # Step 1: Locate the zip file ending with "-GC.zip"
    zip_file_path = None
    for root, _, files in os.walk(data_folder_dir):
        for file in files:
            if file.endswith("-GC.zip"):
                zip_file_path = os.path.join(root, file)
                break
        if zip_file_path:
            break

    if not zip_file_path:
        print("No zip file ending with '-GC.zip' found in the specified directory.")
        return

    print(f"Located zip file: {zip_file_path}")

    # Step 2: Initialize counters for annotations:name and sequence:location
    name_counter = Counter()
    location_counter = Counter()

    # Step 3: Analyze the contents of the zip file
    try:
        with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
            for file_name in zip_ref.namelist():
                # Only process .fusion-data files
                if file_name.endswith(".fusion-data"):
                    with zip_ref.open(file_name) as file:
                        try:
                            # Load JSON and extract fields
                            data = json.load(file)

                            # Count occurrences of "annotations:name"
                            if "annotations" in data and isinstance(data["annotations"], dict):
                                name = data["annotations"].get("name", "No name field found")
                                name_counter[name] += 1

                            # Count occurrences of "sequence:location"
                            if "sequence" in data and isinstance(data["sequence"], dict):
                                location = data["sequence"].get("location", "No location field found")
                                location_counter[location] += 1

                        except json.JSONDecodeError:
                            # Skip files that are not valid JSON
                            print(f"File {file_name} is not a valid JSON format.")

    except Exception as e:
        print(f"An error occurred: {e}")
        return

    # Step 4: Print the variety and occurrences of annotations:name and sequence:location
    print("\nOccurrences of annotations:name:")
    for name, count in name_counter.items():
        print(f"  Name: {name} : Occurrence: {count}")

    print("\nOccurrences of sequence:location:")
    for location, count in location_counter.items():
        print(f"  Location: {location} : Occurrence: {count}")


def update_annotations_in_zip(data_folder_dir: str, annotation_name: str, sequence_location: str) -> None:
    """
    Update annotations in a zip file ending with "-GC.zip" within a given directory.

    Args:
        data_folder_dir (str): Directory containing the data folder.
        annotation_name (str): New value for the "name" field in "annotations".
        sequence_location (str): New value for the "location" field in "sequence".
    """
    # Step 1: Locate the zip file ending with "-GC.zip"
    zip_file_path = None
    for root, _, files in os.walk(data_folder_dir):
        for file in files:
            if file.endswith("-GC.zip"):
                zip_file_path = os.path.join(root, file)
                break
        if zip_file_path:
            break

    if not zip_file_path:
        print("No zip file ending with '-GC.zip' found in the specified directory.")
        return

    print(f"Located zip file: {zip_file_path}")

    # Step 2: Define a temporary directory for unzipping
    temp_dir = os.path.join(data_folder_dir, "temp_unzip")

    try:
        # Ensure the temporary directory is empty before processing
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        os.makedirs(temp_dir)

        # Step 3: Unzip the file
        with zipfile.ZipFile(zip_file_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)

        # Step 4: Update the annotations in all .fusion-data files
        for root, _, files in os.walk(temp_dir):
            for file in files:
                if file.endswith(".fusion-data"):
                    file_path = os.path.join(root, file)
                    with open(file_path, "r", encoding="utf-8") as f:
                        data = json.load(f)

                    # Update the "name" field in "annotations"
                    if "annotations" in data and isinstance(data["annotations"], dict):
                        data["annotations"]["name"] = annotation_name

                    # Update the "location" field in "sequence"
                    if "sequence" in data and isinstance(data["sequence"], dict):
                        data["sequence"]["location"] = sequence_location

                    # Save the updated content back to the file
                    with open(file_path, "w", encoding="utf-8") as f:
                        json.dump(data, f, indent=4)

        # Step 5: Rezip the content back and overwrite the original zip file
        with zipfile.ZipFile(zip_file_path, "w", zipfile.ZIP_DEFLATED) as zip_ref:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    zip_ref.write(file_path, os.path.relpath(file_path, temp_dir))

        print(f"Updated zip file saved at: {zip_file_path}")

    except Exception as e:
        print(f"An error occurred: {e}")

    finally:
        # Step 6: Clean up temporary files and directories
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)
        print("Temporary files cleaned up.")


def gc_points_pruning(dir_folder: str, list_to_delete: List[int]) -> None:
    """
    Prunes files in a GC zip file located in the specified directory by deleting specified files.

    Args:
        dir_folder (str): Path to the directory containing the GC zip file.
        list_to_delete (List[int]): List of file indices (starting from 1) to delete after sorting chronologically.

    Returns:
        None
    """
    # Locate the zip file ending with '-GC.zip'
    gc_zip_path = None
    for file_name in os.listdir(dir_folder):
        if file_name.endswith("-GC.zip"):
            gc_zip_path = os.path.join(dir_folder, file_name)
            break

    if not gc_zip_path:
        raise FileNotFoundError("No zip file ending with '-GC.zip' found in the specified directory.")

    # Create a temporary directory to extract the files
    with tempfile.TemporaryDirectory() as temp_dir:
        # Extract the zip file
        with zipfile.ZipFile(gc_zip_path, "r") as zip_ref:
            zip_ref.extractall(temp_dir)

        # Collect all files and arrange them chronologically
        extracted_files = []
        for root, _, files in os.walk(temp_dir):
            for file in files:
                timestamp_str = file.split("- ")[1].split(".fusion")[0]
                timestamp = datetime.strptime(timestamp_str, "%b %d %Y, %H;%M")
                extracted_files.append((os.path.join(root, file), timestamp))

        # Sort the files based on timestamp
        extracted_files.sort(key=lambda x: x[1])

        # Convert list_to_delete from 1-based to 0-based index
        indices_to_delete = [idx - 1 for idx in list_to_delete]

        # Delete the specified files
        for idx in sorted(indices_to_delete, reverse=True):
            if 0 <= idx < len(extracted_files):
                os.remove(extracted_files[idx][0])

        # Create a new zip file with the same name, ensuring correct file structure
        with zipfile.ZipFile(gc_zip_path, "w", zipfile.ZIP_DEFLATED) as new_zip:
            for root, _, files in os.walk(temp_dir):
                for file in files:
                    abs_path = os.path.join(root, file)
                    rel_path = os.path.relpath(abs_path, temp_dir)
                    new_zip.write(abs_path, rel_path)


def print_metadata_differences(dir_metadatas):
    """
    Prints the variety and frequency of metadata values across multiple Excel files in a folder.

    Args:
        dir_metadatas (str): Path to the folder containing Excel metadata files.
    """
    # Load the first Excel file to get the row names (assuming first column is 'Metadata')
    try:
        first_metadata = os.listdir(dir_metadatas)[0]
        df_metadata = pd.read_excel(os.path.join(dir_metadatas, first_metadata), sheet_name="Metadata")
        row_names = df_metadata["Metadata"]
    except Exception as e:
        print(f"Error loading metadata files or retrieving row names: {e}")
        return

    # Function to normalize the value
    def normalize_value(value):
        try:
            if isinstance(value, pd.Timestamp):  # Check if the value is a datetime
                return value  # Return the datetime as is
            else:
                return float(value)
        except (ValueError, TypeError):
            return str(value)  # Keep as string if it can't be converted to float

    # Function to print variety and frequency
    def print_list_variety(lst):
        count = Counter(lst)
        for value, frequency in count.items():
            print(f"Value: {value}, Frequency: {frequency}")

    # Iterate through each row in the 'Metadata' column
    for row_name in row_names:
        try:
            test = []

            # Iterate through all Excel files and collect values for the current row
            for metadata in os.listdir(dir_metadatas):
                df_metadata = pd.read_excel(os.path.join(dir_metadatas, metadata), sheet_name="Metadata")

                # Find the value corresponding to the current row (row_name)
                value = df_metadata.loc[df_metadata["Metadata"] == row_name, "Value"].values[0]  # Assuming values are in 'Value' column

                # Normalize the value (except for datetime values, which are already in correct format)
                normalized_value = normalize_value(value)

                test.append(normalized_value)

            # Print the variety and frequency for the current row
            print(f"Row name: {row_name}")

            if all(isinstance(v, pd.Timestamp) for v in test):
                for v in test:
                    print(f"Date: {v}")
            else:
                print_list_variety(test)

            print("\n")
        except Exception as e:
            print(f"!!!!!!! There is an issue with row: {row_name}, Error: {e}")
            print("\n")


def process_pseudo_list(metadata_str: str) -> list:
    """
    Process a pseudo-list string (found in the metadata) into a list or a list of lists.

    This function takes a pseudo-list string (e.g., "[a,b,c],[d,e,f]") and converts it into
    a Python list. If the input string represents nested pseudo-lists, it returns a list of lists.
    The function ensures proper formatting of the input string, even if it lacks spaces between
    elements or sublists.

    Args:
        metadata_str (str): A string representing a pseudo-list, potentially containing nested
                            pseudo-lists separated by "],[" or similar patterns.

    Returns:
        list: A processed list or a list of lists, depending on the structure of the input string.
    """
    ls_result = []
    # format the input string into a correct format (in case the user do not put space between the commas between different psudolist)
    if "],[" in metadata_str:
        metadata_str.replace("],[", "], [")
    # format the input string into a correct format (in case the user do not put space between the commas between different psudolist)
    if "][" in metadata_str:
        metadata_str.replace("][", "], [")

    list_to_process = metadata_str.split("], [")
    for item in list_to_process:
        ls_extracted_item = [info for info in item.strip("[]").split(",")]
        ls_result.append(ls_extracted_item)
    return ls_result


def get_element_from_pseudolist(nested_list: list, indices: list) -> any:
    """
    Retrieve an element from a one- or two-level nested list based on a single parameter.

    Args:
        nested_list (list): A list or a list of lists to retrieve an element from. This is the result from process_pseudo_list function.
        indices (list): A list of one or two indices specifying the path to the desired element.

    Returns:
        any: The element at the specified path of indices.

    Raises:
        ValueError: If the indices list is empty or has more than two elements.
    """
    if not indices or len(indices) > 2:
        raise ValueError("Indices list must have one or two elements only.")

    if len(indices) == 1:
        # Single index: Access the first level
        return nested_list[indices[0]]
    elif len(indices) == 2:
        # Two indices: Access the second level
        return nested_list[indices[0]][indices[1]]


def inspect_gc_folder(dir_gc_zip: str) -> None:
    """
    Inspect the contents of a GC folder and print the names of all files and subfolders.

    Args:
        dir_gc_zip (str): Path to the GC folder.

    Returns:
        None
    """
    # Sort out path and folder name for extracting the zip file
    dir_gc_zip = Path(dir_gc_zip)
    dir_base_folder = dir_gc_zip.parent
    dir_folder_name = dir_gc_zip.stem

    # Extract the zip file to a temporary directory
    with zipfile.ZipFile(dir_gc_zip, "r") as zip_ref:
        temp_dir = dir_base_folder / f"{dir_folder_name}_temp"
        zip_ref.extractall(temp_dir)

    # Try to call yadg's fusion data extraction function.
    ls_issue_files = []
    for fusion_file in temp_dir.iterdir():
        try:
            fusion_json.extract(fn=fusion_file, encoding="utf-8", timezone="UTC")
        except:
            ls_issue_files.append(fusion_file.name)

    # Report the issue
    if not ls_issue_files:
        print("All files processed successfully.")
    else:
        for file in ls_issue_files:
            print(f"File {file} has an issue")

    # Delete the temporary directory
    shutil.rmtree(temp_dir)
