"""
This module contains function desgined to handle ontology xlsx file generation and upload
"""

import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

from . import keller, pathfolio


def update_metadata_value(file_path, metadata, input_value, sheet_name="Schema"):
    """
    Update the value of a specified metadata key in an Excel sheet.

    Args:
        file_path (str): Path to the Excel file.
        metadata (str): The metadata key to search for.
        input_value (str): The new value to set.
        sheet_name (str): Name of the sheet to search in (default is "Schema").
    """
    try:
        # Load the workbook and select the specified sheet
        workbook = load_workbook(file_path)
        if sheet_name not in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' not found in the workbook.")
            return

        sheet = workbook[sheet_name]

        # Find the row with the specified metadata and update the corresponding value
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
            metadata_cell, value_cell = row
            if metadata_cell.value == metadata:
                value_cell.value = input_value
                print(f"Updated metadata '{metadata}' with value '{input_value}'.")
                break
        else:
            print(f"Metadata '{metadata}' not found in sheet '{sheet_name}'.")

        # Save the changes to the Excel file
        workbook.save(file_path)

    except Exception as e:
        print(f"An error occurred: {e}")


def gen_metadata_xlsx(
    dir_json: str,
    dir_template: str = r"K:\Aurora\nukorn_PREMISE_space\Battinfo_template.xlsx",
) -> None:
    """
    Generate a metadata Excel file for a specific experiment based on a template.

    After generating this metadata file, the script will automatically extract the metadata information form the analyzed json file and automatcailly update the
    metadata values in the new Excel file.
    This Excel file will then be used as a metadata excel file used in generating a corresponding ontologized JSON-LD file.

    Args:
        dir_json (str): The path to the analyzed JSON file.
        dir_template (str): The path to the template Excel file. Defaults to
                            'K:\\Aurora\\nukorn_PREMISE_space\\Battinfo_template.xlsx'.


    Returns:
        None: Creates a new Excel file in the backup directory with the experiment name as part of the file name.
    """
    dir_json = Path(dir_json)
    # Extract the experiment name from the analyzed JSON file path
    experiment_name = dir_json.stem.split(".")[1]
    # Get the name of the new xlsx file
    new_xlsx_name = f"{experiment_name}_automated_extract_metadata.xlsx"
    dir_new_xlsx = dir_json.parent / new_xlsx_name
    # Copy the template file
    shutil.copy(dir_template, dir_new_xlsx)

    # Update the experiment name in the new Excel file
    dict_metadata = curate_metadata_dict(dir_json)
    for key, value in dict_metadata.items():
        update_metadata_value(dir_new_xlsx, key, value)


def curate_metadata_dict(dir_json: str) -> Dict[str, str]:
    """
    Generates a metadata dictionary by extracting relevant information from a JSON file.

    This metadata is used to populate an Excel file that will later generate an
    ontologized JSON-LD file for further use.

    Args:
        dir_json (str): The file path to the JSON file that contains the analyzed metadata.

    Returns:
        Dict[str, str]: A dictionary containing metadata items as keys and their corresponding values.

    Raises:
        ValueError: If the operator's short name is not recognized or if the date format in the
                    'Cell ID' field does not follow the expected 'yymmdd' format.
    """
    dict_metadata = {}

    # Extract metadata from the analyzed json file.
    for key, value in pathfolio.dict_excel_to_json.items():
        dict_metadata[key] = keller.get_metadata_from_json(dir_json, value)

    # Extracting operator name
    user_short_name = dict_metadata["Cell ID"].split("_")[1]
    user_full_name = pathfolio.user_mapping.get(user_short_name)

    if user_full_name is None:
        raise ValueError(
            f"{user_short_name} is not recognized, please consult with Graham Kimbell. He will be delighted to assist you."
        )

    dict_metadata["Scientist/technician/operator"] = user_full_name

    # Extracting the date of the experiment
    try:
        unformatted_date_string = dict_metadata["Cell ID"].split("_")[0]
        parsed_date = datetime.strptime(unformatted_date_string, "%y%m%d")
        formatted_date_string = parsed_date.strftime("%d/%m/%Y")
        dict_metadata["Date of cell assembly"] = formatted_date_string
    except:
        raise ValueError(
            f"Date extarcted from the cell ID {unformatted_date_string} is not in the correct format of yymmdd, please check the format"
        )
    return dict_metadata
