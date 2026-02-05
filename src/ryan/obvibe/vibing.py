"""
This is the main module for this repository
"""

import shutil
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook

from . import keller, oh_my_ontology, pathfolio


class Identifiers:
    """
    Class object help with the identification of space, project and experiment in openBIS.
    """

    def __init__(self, space_code: str, project_code: str, experiment_code: str) -> None:
        self.space_code = space_code
        self.project_code = project_code
        self.experiment_code = experiment_code

    @property
    def space_identifier(self) -> str:
        return self.space_code

    @property
    def project_identifier(self) -> str:
        return f"/{self.space_identifier}/{self.project_code}"

    @property
    def experiment_identifier(self) -> str:
        return f"{self.project_identifier}/{self.experiment_code}"


class Dataset:
    """
    Class made to facilitate dataset upload
    """

    def __init__(self, openbis_instance, ident: Identifiers, dataset_type=None, upload_data=None) -> None:
        self.ob = openbis_instance
        self.ident = ident
        self.type = dataset_type
        self.data = upload_data
        self.experiment = self.ident.experiment_identifier.upper()  # Use the provided Identifiers instance

    def upload_dataset(self):
        """
        Upload the dataset to the openBIS
        """
        self.ob.new_dataset(type=self.type, experiment=self.experiment, file=self.data).save()


def push_exp(
    dir_pat: str,
    dir_folder: str,
    dict_mapping: dict = pathfolio.metadata_mapping,
    space_code: str = "TEST_SPACE_PYBIS",
    project_code: str = "TEST_UPLOAD",
    experiment_type: str = "Battery_Premise2",
) -> None:
    """
    Pushes experimental data and metadata from a local folder to an openBIS instance.

    Args:
        dir_pat (str): Path to the openBIS PAT file (personal access token).
        dir_folder (str): Path to the directory containing the experimental data files.
        dict_mapping (dict, optional): A mapping dictionary defining openBIS codes and JSON paths for metadata extraction. Defaults to `pathfolio.metadata_mapping`.
        space_code (str, optional): The openBIS space code where the experiment will be created. Defaults to 'TEST_SPACE_PYBIS'.
        project_code (str, optional): The openBIS project code where the experiment will be created. Defaults to 'TEST_UPLOAD'.
        experiment_type (str, optional): The type of experiment to be created in openBIS. Defaults to 'Battery_Premise2'.

    Raises:
        ValueError: If there is not exactly one JSON file in the specified folder.
        ValueError: If the JSON file name does not follow the required naming convention.
        ValueError: If there is not exactly one raw HDF5 file in the specified folder.

    Returns:
        None
    """
    dir_folder = Path(dir_folder)
    ob = keller.get_openbis_obj(dir_pat)

    list_json = [
        file for file in dir_folder.iterdir() if file.suffix == ".json" and not file.stem.startswith("ontologized")
    ]
    if len(list_json) != 1:
        raise ValueError("There should be exactly one json file in the folder")

    name_json = list_json[0].name
    dir_json = dir_folder / name_json

    if len(dir_json.name.split(".")) != 3:
        raise ValueError("Not recognized json file name. The recognized file name is cycle.experiment_code.json")

    exp_name = dir_json.stem.split(".")[1]  # Extract the experiment name from the json file name
    ident = Identifiers(space_code, project_code, experiment_code=exp_name)

    # Create new experiment in the predefined space and project.
    exp = ob.new_experiment(code=ident.experiment_code, type=experiment_type, project=ident.project_identifier)

    # Iterate through list of metadata from json file and upload them to the experiment.
    for item in dict_mapping:
        try:
            openbis_code = item["openbis_code"]
            json_path = item["json_path"]
            print(f"Uploading metadata for {openbis_code} from {json_path}")
            exp.p[openbis_code] = keller.get_metadata_from_json(dir_json, json_path)
            exp.save()
        except Exception as e:
            print(f"Error uploading metadata for {openbis_code} from {json_path}")
            print(f"The error message is: {e} \n")
            continue

    # Upload the dataset

    # Analyzed data
    ds_analyzed_json = Dataset(ob, ident=ident)
    ds_analyzed_json.type = "premise_cucumber_analyzed_battery_data"
    ds_analyzed_json.data = dir_json
    ds_analyzed_json.upload_dataset()

    # Raw data
    list_raw_data = [
        file for file in dir_folder.iterdir() if file.suffix == ".h5" and file.stem.split(".")[1] == exp_name
    ]
    if len(list_raw_data) != 1:
        raise ValueError("There should be exactly one raw_h5 file in the folder")

    dir_raw_json = list_raw_data[0]
    ds_raw_data = Dataset(ob, ident=ident)
    ds_raw_data.type = "premise_cucumber_raw_battery_data"
    ds_raw_data.data = dir_raw_json
    ds_raw_data.upload_dataset()

    # Create the automated_extract_metadata.xlsx file
    oh_my_ontology.gen_metadata_xlsx(dir_json)
    source_file = dir_folder / f"{exp_name}_automated_extract_metadata.xlsx"
    dest_file = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    print(f"Copying {source_file} to {dest_file}")
    shutil.copy(source_file, dest_file)

    # Check if there is already a custom Excel file for the experiment. If so, create JSON-LD from it.
    custom_metadata_files = [file for file in dir_folder.iterdir() if file.name.endswith("custom_metadata.xlsx")]
    if custom_metadata_files:
        custom_metadata = custom_metadata_files[0]

        # Load both Excel files
        merged_wb = load_workbook(dest_file)
        custom_wb = load_workbook(custom_metadata)

        # Select the "Schema" sheet from both workbooks
        merged_sheet = merged_wb["Schema"]
        custom_sheet = custom_wb["Schema"]

        # Find the "Value" column index in the "Schema" sheet
        header_row = 1  # Assuming headers are in the first row
        value_column_index = None

        for col in range(1, custom_sheet.max_column + 1):
            if custom_sheet.cell(row=header_row, column=col).value == "Value":
                value_column_index = col
                break

        if value_column_index is None:
            raise ValueError("Column 'Value' not found in the 'Schema' sheet.")

        # Loop through rows in the "Value" column of the custom metadata
        for row in range(header_row + 1, custom_sheet.max_row + 1):  # Skip the header row
            custom_value = custom_sheet.cell(row=row, column=value_column_index).value
            if custom_value:  # Skip if the cell is empty or None
                # Write the custom value into the corresponding row of the merged metadata
                merged_sheet.cell(row=row, column=value_column_index).value = custom_value

        # Save the updated merged metadata workbook
        merged_wb.save(dest_file)

    # Upload the metadata Excel file to the openBIS
    dir_metadata_excel = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    ds_metadata_excel = Dataset(ob, ident=ident)
    ds_metadata_excel.type = "premise_excel_for_ontology"
    ds_metadata_excel.data = dir_metadata_excel
    ds_metadata_excel.upload_dataset()

    # Generate the ontologized JSON-LD file
    jsonld_filename = f"ontologized_{exp_name}.json"
    dir_xlsx = dir_folder / f"{exp_name}_merged_metadata.xlsx"
    oh_my_ontology.gen_jsonld(dir_xlsx, jsonld_filename)

    # Upload the ontologized JSON-LD file to the openBIS
    dir_jsonld = dir_folder / jsonld_filename
    ds_jsonld = Dataset(ob, ident=ident)
    ds_jsonld.type = "premise_jsonld"
    ds_jsonld.data = dir_jsonld
    ds_jsonld.upload_dataset()


def gen_pmd_associate_battery_obj(dir_pat: str, battery_name: str) -> dict:
    """
    Generate battery objects in openBIS and associate them with each other.

    This function creates the battery objects, following a discussion with Edan (late march 2024).
    The function also returns a dictionary containing the permIDs of the created objects.

    Args:
        dir_pat (str): Path to the openBIS PAT file (personal access token).
        battery_name(str): Name of the battery (assembled and test).


    Returns:
        dict: A dictionary containing perID of the created objects.
    """
    dict_components = {}
    ob = keller.get_openbis_obj(dir_pat)

    # Create the battery experiment
    new_battery_exp = ob.new_experiment(
        code=battery_name,
        type="default_experiment",
        project="/LAB501_NUKORN.PLAINPAN_AT_EMPA.CH/Pmd_Test_project",
    )
    new_battery_exp.save()
    new_battery_exp_permID = new_battery_exp.permId
    dict_components["battery_experiment"] = new_battery_exp_permID

    # Create components to be assembled.
    positive_electrode = ob.new_object(
        type="Pmd_Positive_Electrode",
        space="/LAB501_MATERIALS",
        project="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS",
        collection="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS/PMD_BATTERY_COMPONENTS_EXP_1",
    )
    positive_electrode.save()
    positive_electrode_permID = positive_electrode.permId
    dict_components["positive_electrode"] = positive_electrode_permID

    negative_electrode = ob.new_object(
        type="Pmd_Negative_Electrode",
        space="/LAB501_MATERIALS",
        project="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS",
        collection="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS/PMD_BATTERY_COMPONENTS_EXP_2",
    )
    negative_electrode.save()
    negative_electrode_permID = negative_electrode.permId
    dict_components["negative_electrode"] = negative_electrode_permID

    electrolyte = ob.new_object(
        type="Pmd_electrolyte",
        space="/LAB501_MATERIALS",
        project="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS",
        collection="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS/PMD_BATTERY_COMPONENTS_EXP_3",
    )
    electrolyte.save()
    electrolyte_permID = electrolyte.permId
    dict_components["electrolyte"] = electrolyte_permID

    separator = ob.new_object(
        type="Pmd_Separator",
        space="/LAB501_MATERIALS",
        project="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS",
        collection="/LAB501_MATERIALS/PMD_BATTERY_COMPONENTS/PMD_BATTERY_COMPONENTS_EXP_4",
    )
    separator.save()
    separator_permID = separator.permId
    dict_components["separator"] = separator_permID

    # Create assmbled battery and the cycled battery.

    assembled_battery = ob.new_object(
        type="Pmd_assembled_battery",
        space="/LAB501_PMD_BATTERY_INVENTORY",
        project="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2",
        collection="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2/PMD_TEST_PROJECT2_EXP_1",
    )
    assembled_battery.save()
    assembled_battery_permID = assembled_battery.permId
    dict_components["assembled_battery"] = assembled_battery_permID

    formation_cycled_battery = ob.new_object(
        type="Pmd_cycled_battery",
        space="/LAB501_PMD_BATTERY_INVENTORY",
        project="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2",
        collection="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2/PMD_TEST_PROJECT2_EXP_2",
    )
    formation_cycled_battery.save()
    formation_cycled_battery_permID = formation_cycled_battery.permId
    dict_components["formation_cycled_battery"] = formation_cycled_battery_permID

    long_term_cycled_battery = ob.new_object(
        type="Pmd_cycled_battery",
        space="/LAB501_PMD_BATTERY_INVENTORY",
        project="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2",
        collection="/LAB501_PMD_BATTERY_INVENTORY/PMD_TEST_PROJECT2/PMD_TEST_PROJECT2_EXP_2",
    )
    long_term_cycled_battery.save()
    long_term_cycled_battery_permID = long_term_cycled_battery.permId
    dict_components["long_term_cycled_battery"] = long_term_cycled_battery_permID

    # Create protocols (assembly and cycling protocols).
    path = PurePosixPath(new_battery_exp.identifier)
    _space = f"/{path.parts[1]}"
    _project = f"/{'/'.join(path.parts[1:3])}"
    _experiment = str(path)

    battery_assembly = ob.new_object(
        type="Pmd_Battery_Assembly", space=_space, project=_project, experiment=_experiment
    )
    battery_assembly.save()
    battery_assembly_permID = battery_assembly.permId
    dict_components["battery_assembly"] = battery_assembly_permID

    formation_cycle_protocol = ob.new_object(
        type="PMD_BATTERY_CYCLING_PROTOCOL", space=_space, project=_project, experiment=_experiment
    )
    formation_cycle_protocol.save()
    formation_cycle_protocol_permID = formation_cycle_protocol.permId
    dict_components["formation_cycle_protocol"] = formation_cycle_protocol_permID

    long_term_cycle_protocol = ob.new_object(
        type="PMD_BATTERY_CYCLING_PROTOCOL", space=_space, project=_project, experiment=_experiment
    )
    long_term_cycle_protocol.save()
    long_term_cycle_protocol_permID = long_term_cycle_protocol.permId
    dict_components["long_term_cycle_protocol"] = long_term_cycle_protocol_permID

    # Make parent-child relation to all created battery objects
    battery_assembly.add_parents(positive_electrode)
    battery_assembly.add_parents(negative_electrode)
    battery_assembly.add_parents(electrolyte)
    battery_assembly.add_parents(separator)
    battery_assembly.save()
    assembled_battery.add_parents(battery_assembly)
    assembled_battery.save()
    formation_cycle_protocol.add_parents(assembled_battery)
    formation_cycle_protocol.save()
    formation_cycled_battery.add_parents(formation_cycle_protocol)
    formation_cycled_battery.save()
    long_term_cycle_protocol.add_parents(formation_cycled_battery)
    long_term_cycle_protocol.save()
    long_term_cycled_battery.add_parents(long_term_cycle_protocol)
    long_term_cycled_battery.save()

    return dict_components


def fill_in_pmd_battery_instance(dir_data_folder: Path, dict_permid_battery_experiment: dict, dir_pat: str) -> None:
    """
    Fill in the PMD battery instance with the relevant information and associated dataset

    At the current state, this function will rely on the data from Aurora tools.

    Args:
        dir_data_folder (pathlib.Path): Path to the directory containing the experimental data files.
        dict_permid_battery_experiment (dict): A dictionary containing permIDs of the created battery objects.
                                               This is generated from gen_pmd_associate_battery_obj function.
        dir_pat (str): Path to the openBIS PAT file (personal access token).

    returens:
        None
    """
    # Make sure the input directory is a Path object
    if not isinstance(dir_data_folder, Path):
        if isinstance(dir_data_folder, str):
            dir_data_folder = Path(dir_data_folder)
        else:
            raise TypeError("dir_data should be a string or pathlib.Path object")

    # Locate the raw HDF5 file in the directory
    matchs_full_h5_dir = list(dir_data_folder.glob("full*.h5"))

    if len(matchs_full_h5_dir) == 0:
        raise FileNotFoundError("No raw HDF5 files found in the specified directory.")
    elif len(matchs_full_h5_dir) > 1:
        raise ValueError("More than one raw HDF5 file found in the specified directory. Please check the directory.")
    dir_raw_h5 = matchs_full_h5_dir[0]

    # #Generate associated cycling protocol and battery instance
    keller.gen_protocol_data(dir_raw_h5)

    ob = keller.get_openbis_obj(dir_pat)
    dict_dataset = {
        "formation_cycle_protocol": dir_data_folder / "formation_cycle_protocol.h5",
        "formation_cycled_battery": dir_data_folder / "formation_cycle_data.h5",
        "long_term_cycle_protocol": dir_data_folder / "long_term_cycle_protocol.h5",
        "long_term_cycled_battery": dir_data_folder / "long_term_cycle_data.h5",
    }

    # Upload the datasets to openBIS
    for key, value in dict_dataset.items():
        keller.upload_dataset(ob, value, dict_permid_battery_experiment[key], sample_type="pmd_general_dataset")

    # Write a comment for each object

    dict_comments = {
        "positive_electrode": "place holder for information for Positive Electrode",
        "negative_electrode": "place holder for information for Negative Electrode",
        "electrolyte": "place holder for information for Electrolyte",
        "separator": "place holder for information for Separator",
        "assembled_battery": "place holder for information for Assembled Battery",
        "formation_cycled_battery": "place holder for information for Formation Cycled Battery",
        "long_term_cycled_battery": "place holder for information for Long Term Cycled Battery",
        "battery_assembly": "place holder for information for Battery Assembly",
        "formation_cycle_protocol": "place holder for information for Formation Cycle Protocol",
        "long_term_cycle_protocol": "place holder for information for Long Term Cycle Protocol",
    }

    for key, value in dict_comments.items():
        obj = ob.get_object(dict_permid_battery_experiment[key])
        obj.p["notes"] = value
        obj.save()


def gen_pmd_battery_instance(dir_pat: str, battery_name: str, dir_data_folder: Path) -> None:
    """
    Generate PMD battery instance and fill in the relevant information.

    Args:
        dir_pat (str): Path to the openBIS PAT file (personal access token).
        battery_name (str): Name of the battery (assembled and test).
        dir_data_folder (pathlib.Path): Path to the directory containing the experimental data files.

    Returns:
        None
    """
    dict_component_permid = gen_pmd_associate_battery_obj(dir_pat, battery_name)
    fill_in_pmd_battery_instance(dir_data_folder, dict_component_permid, dir_pat)
